# coding = utf - 8

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from appium import webdriver
import openpyxl
from openpyxl import load_workbook
import os
import time

adrress='phone.xlsx'
wb = openpyxl.load_workbook(adrress)     #打开excel文件
# print(wb.sheetnames)                 #获取工作簿所有表格
sheet = wb["Sheet1"]  #获取工作表
phone=[]
password=[]
for i in sheet["A2":"B"+str(sheet.max_row)]:
    phone.append(i[0].value)
    password.append(i[1].value)
print(phone)
print(password)


'''判断toast信息'''
def find_toast(toast):
    try:
        ele = WebDriverWait(driver, 5).until(expected_conditions.presence_of_element_located((By.XPATH, './/*[contains(@text,'+'\''+toast+'\''+')]')))
        print(toast)
        return True
    except:
        print('aaa')
        return False
'''返回可见元素'''
def visible_ele(driver, identifyBy, c):
    try:
        if(identifyBy == "id"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.ID, c)))
            print(element.text)
        elif(identifyBy == "xpath"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH,c)))
        elif(identifyBy == "class_name"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CLASS_NAME,c)))
        elif(identifyBy == "tag_name"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.TAG_NAME,c)))
        elif(identifyBy == "css_selector"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR,c)))
        elif(identifyBy == "link_text"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.LINK_TEXT,c)))
        elif(identifyBy == "name"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.NAME,c)))
        elif(identifyBy == "partial_link_text"):
            element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.PARTIAL_LINK_TEXT,c)))
    except(NoSuchElementException, e):
        print("未找到元素："+c)
    finally:
        return element
#判断界面元素是否存在
def isElement(driver,identifyBy,c):
    '''
    Determine whether elements exist
    Usage:
    isElement(By.XPATH,"//a")
    '''
    time.sleep(1)
    flag=None
    try:
        if identifyBy == "id":
            #self.driver.implicitly_wait(60)
            driver.find_element_by_id(c)
        elif identifyBy == "xpath":
            #self.driver.implicitly_wait(60)
            driver.find_element_by_xpath(c)
        elif identifyBy == "class":
            driver.find_element_by_class_name(c)
        elif identifyBy == "link text":
            driver.find_element_by_link_text(c)
        elif identifyBy == "partial link text":
            driver.find_element_by_partial_link_text(c)
        elif identifyBy == "name":
            driver.find_element_by_name(c)
        elif identifyBy == "tag name":
            driver.find_element_by_tag_name(c)
        elif identifyBy == "css selector":
            driver.find_element_by_css_selector(c)
        flag = True
    except(NoSuchElementException, e):
        print('teyey')
        flag = False
    finally:
        return flag


#左滑
def getSize():
    x = driver.get_window_size()['width']
    y = driver.get_window_size()['height']
    return (x, y)
# 屏幕向左滑动
def swipLeft(t):  #左滑方法，调用此方法传入t参数即可
    l = getSize()
    x1 = int(l[0] * 0.75)
    y1 = int(l[1] * 0.5)
    x2 = int(l[0] * 0.05)
    driver.swipe(x1, y1, x2, y1, t)

# #    获取手机信息
desired_caps = {
    'platformName': 'Android',
    'deviceName': 'YHQGORTWJV5SAQOV',  #'77a3c838'
    'platformVersion': '8.1.0',
    'appPackage': 'com.bianla.tangba',     # apk包名
    'appActivity': 'com.bianla.tangba.activity.StartActivity',      # apk的launcherActivity
    'noReset': 'true',
    # 'unicodeKeyboard': 'true',
    # 'resetKeyboard': 'true',
    'automationName':'uiautomator2'
    # 'automationName':'Selendroid'
}
driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', desired_caps)

# time.sleep(2)
# swipLeft(1000)
# time.sleep(1)
# swipLeft(1000)

#driver.find_element_by_id("com.bianla.tangba:id/btn_skip").click()  #点击导航页“立即体验”按钮
visible_ele(driver,id, "com.bianla.tangba:id/btn_skip").click()


for i in range(0,sheet.max_row-1):
    time.sleep(2)
    driver.find_element_by_id("com.bianla.tangba:id/btnEnter").click()  # 点击“注册登陆”按钮
    time.sleep(1)
    driver.find_element_by_id("com.bianla.tangba:id/etPhone").send_keys(phone[i])  #输入账号 来自表格
    time.sleep(2)
    driver.find_element_by_id("com.bianla.tangba:id/btnNext").click()  #点击“下一步”
    time.sleep(3)
    # 正则匹配手机号
    if(len(phone[i])<11):
        print(len(phone[i]))
        if (find_toast('请输入有效的手机号')):
            sheet['C' + str(i + 2)] = '请输入有效的手机号'
            wb.save(adrress)
            driver.back()
    elif (find_toast('手机号码不合法')):
        sheet['C' + str(i + 2)] = '手机号码不合法'
        wb.save(adrress)
        driver.back()

    elif (isElement(driver,'id','com.bianla.tangba:id/btnEnter')):   #如果是登录按钮，进入以下判断

        driver.find_element_by_id("com.bianla.tangba:id/etPassword").send_keys(password[i])  #输入密码 来自表格
        driver.find_element_by_id("com.bianla.tangba:id/btnEnter").click()  #点击“登陆”

        if (find_toast('用户名或密码错误')):
            sheet['C' + str(i + 2)] = '用户名或密码错误'
            wb.save(adrress)
            driver.back()
            driver.back()

        else:
            driver.implicitly_wait(20)
            #driver.find_element_by_id("com.bianla.tangba:id/dialog_urine_ketone_iv_close").click()  # 点击“设置通知弹框”的关闭
            sheet['C'+str(i+2)]= '运行成功啦'
            wb.save(adrress)
            driver.implicitly_wait(20)
            driver.background_app(0.5)
            time.sleep(2)
            driver.find_element_by_id("com.bianla.tangba:id/action_mine").click()  #点击“我的”
            driver.find_element_by_id("com.bianla.tangba:id/iv_add").click()  #点击右上角“设置”按钮
            driver.find_element_by_xpath("//android.support.v7.widget.RecyclerView[@resource-id=\"com.bianla.tangba:id/recycler_view\"]/android.widget.RelativeLayout[4]").click()
            time.sleep(2)
            driver.find_element_by_id("com.bianla.tangba:id/confirm").click()  #点击“确定退出”

    elif (isElement(driver,'id','com.bianla.tangba:id/btnComplete')) :  #注册进入以下步骤
        print('zhuce')
        driver.find_element_by_id("com.bianla.tangba:id/etVCode").send_keys('1111')
        driver.find_element_by_id("com.bianla.tangba:id/etPassword").send_keys(password[i])
        driver.find_element_by_id("com.bianla.tangba:id/btnComplete").click()  # 点击“完成”按钮
        driver.find_element_by_id("com.bianla.tangba:id/btn_next").click()  # 点击“下一步”按钮
        driver.find_element_by_id("com.bianla.tangba:id/btn_complete").click()  # 点击“完成”按钮
        # driver.find_element_by_id("com.bianla.tangba:id/dialog_urine_ketone_iv_close").click()  # 点击“设置通知弹框”的关闭
        sheet['C'+str(i+2)]= '注册成功'
        wb.save(adrress)
        driver.implicitly_wait(20)
        driver.background_app(0.5)
        driver.find_element_by_id("com.bianla.tangba:id/action_mine").click()  #点击“我的”
        driver.find_element_by_id("com.bianla.tangba:id/iv_add").click()  #点击右上角“设置”按钮
        driver.find_element_by_xpath("//android.support.v7.widget.RecyclerView[@resource-id=\"com.bianla.tangba:id/recycler_view\"]/android.widget.RelativeLayout[4]").click()
        time.sleep(2)
        driver.find_element_by_id("com.bianla.tangba:id/confirm").click()  #点击“确定退出”




